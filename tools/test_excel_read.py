#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test para verificar la lectura del Excel y los precios
"""

import openpyxl

def test_excel():
    file_path = '/home/javier/fsf-pruebaimport/fs-framework/test/tarifa-convertida.xlsx'
    
    # Cargar el archivo
    print("=" * 60)
    print("TEST DE LECTURA DE EXCEL")
    print("=" * 60)
    
    # Probar con data_only=True
    print("\n1. Cargando con data_only=True, read_only=False...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    sheet = wb[wb.sheetnames[0]]
    
    print(f"   Hoja: {wb.sheetnames[0]}")
    print(f"   Total hojas: {len(wb.sheetnames)}")
    
    # Leer encabezados
    print("\n2. Encabezados (fila 1):")
    headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for i, cell in enumerate(headers):
        if cell:
            print(f"   Columna {i} ({chr(65+i)}): '{cell}'")
    
    # Leer primeras filas
    print("\n3. Primeras 15 filas de datos:")
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True)):
        row_list = list(row)
        # Mostrar solo columnas con datos
        row_str = " | ".join([f"{chr(65+i)}:{v}" for i, v in enumerate(row_list[:8]) if v])
        print(f"   Fila {row_idx}: {row_str}")
    
    # Buscar específicamente los artículos con precios
    print("\n4. Buscando filas con ID numérico (artículos):")
    count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=50, values_only=True)):
        row_list = list(row)
        id_col = str(row_list[0]).strip() if row_list[0] else ''
        
        if id_col.isdigit():
            count += 1
            print(f"   Fila {row_idx+2}:")
            print(f"      ID (col A): {row_list[0]}")
            print(f"      SKU (col B): {row_list[1] if len(row_list) > 1 else 'N/A'}")
            print(f"      Código (col C): {row_list[2] if len(row_list) > 2 else 'N/A'}")
            print(f"      Desc ES (col D): {row_list[3] if len(row_list) > 3 else 'N/A'}")
            print(f"      Desc EN (col E): {row_list[4] if len(row_list) > 4 else 'N/A'}")
            print(f"      Biblioteca (col F): {row_list[5] if len(row_list) > 5 else 'N/A'}")
            print(f"      Col G: {row_list[6] if len(row_list) > 6 else 'N/A'}")
            print(f"      Col H: {row_list[7] if len(row_list) > 7 else 'N/A'}")
            print(f"      PVP (col I): {row_list[8] if len(row_list) > 8 else 'N/A'}")
            print(f"      NUEVO (col J): {row_list[9] if len(row_list) > 9 else 'N/A'}")
            print()
            
            if count >= 5:
                break
    
    wb.close()
    
    # Probar también sin data_only para ver si hay fórmulas
    print("\n5. Verificando si hay fórmulas (data_only=False):")
    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    sheet2 = wb2[wb2.sheetnames[0]]
    
    count = 0
    for row_idx, row in enumerate(sheet2.iter_rows(min_row=2, max_row=50)):
        id_cell = row[0]
        if id_cell.value and str(id_cell.value).strip().isdigit():
            count += 1
            pvp_cell = row[8] if len(row) > 8 else None  # Columna I (índice 8)
            if pvp_cell:
                print(f"   Fila {row_idx+2}, PVP:")
                print(f"      Valor: {pvp_cell.value}")
                print(f"      Tipo: {type(pvp_cell.value)}")
                if hasattr(pvp_cell, 'data_type'):
                    print(f"      Data type: {pvp_cell.data_type}")
            if count >= 3:
                break
    
    wb2.close()
    print("\n" + "=" * 60)
    print("TEST COMPLETADO")
    print("=" * 60)

if __name__ == "__main__":
    test_excel()
