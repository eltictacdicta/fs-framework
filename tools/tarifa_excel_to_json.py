#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor de Tarifa Excel a JSON
================================
Esta aplicación convierte archivos de tarifa en formato Excel a JSON
para su importación en el sistema FSFramework.

Estructura esperada del Excel:
- Hoja: "TARIFA COMPLETA"
- Columnas: SKU | Código | Descripción | Description | PVP 2025
- Categorías: "1. ENFRIADORES" (número + punto + texto en mayúsculas)
- Subcategorías: "1.1 Refrigerador de botellas" (N.N + texto)
- Artículos: Filas con SKU en columna A
- Opcionales: Filas después de "OPCIONAL"

Autor: FSFramework Tools
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import re
import os
import uuid
import hashlib
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None


class TarifaConverter:
    """Clase principal para convertir archivos de tarifa Excel a JSON."""
    
    def __init__(self):
        self.data = None
        self.result = {
            'metadata': {
                'generated_at': None,
                'source_file': None,
                'total_categories': 0,
                'total_subcategories': 0,
                'total_articles': 0,
                'total_optionals': 0
            },
            'familias': [],
            'articulos': [],
            'opcionales': [],
            'relaciones_articulo_opcional': []
        }
        # Mapeo de capítulo a código hash para mantener consistencia
        self.capitulo_to_codigo = {}
        
    def load_excel(self, file_path):
        """Carga el archivo Excel y extrae los datos de la hoja TARIFA COMPLETA."""
        if openpyxl is None and pd is None:
            raise ImportError("Se requiere openpyxl o pandas. Instala con: pip install openpyxl pandas")
        
        self.result['metadata']['source_file'] = os.path.basename(file_path)
        self.result['metadata']['generated_at'] = datetime.now().isoformat()
        
        # Intentar con openpyxl primero (más control sobre la lectura)
        if openpyxl:
            return self._load_with_openpyxl(file_path)
        else:
            return self._load_with_pandas(file_path)
    
    def _load_with_openpyxl(self, file_path):
        """Carga el Excel usando openpyxl con data_only=True para leer valores de fórmulas."""
        try:
            # data_only=True lee los valores calculados, no las fórmulas
            # NOTA: read_only=False es necesario para que data_only funcione correctamente
            # con celdas calculadas. El archivo debe haber sido guardado con Excel para
            # que los valores calculados estén disponibles.
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
            
            # Usar la primera hoja del Excel
            sheet_name_found = wb.sheetnames[0]
            sheet = wb[sheet_name_found]
            
            # Convertir a lista de listas
            self.data = []
            for row in sheet.iter_rows(values_only=True):
                # Convertir None a string vacío y todo a string
                row_data = [str(cell) if cell is not None else '' for cell in row]
                self.data.append(row_data)
            
            wb.close()
            return len(self.data), sheet_name_found
            
        except InvalidFileException as e:
            raise ValueError(f"Archivo Excel inválido: {str(e)}")
        except Exception as e:
            raise ValueError(f"Error al leer el archivo: {str(e)}")
    
    def _load_with_pandas(self, file_path):
        """Carga el Excel usando pandas como alternativa."""
        try:
            # Leer nombres de hojas
            xl = pd.ExcelFile(file_path, engine='openpyxl')
            
            # Usar la primera hoja del Excel
            sheet_name = xl.sheet_names[0]
            
            # Leer la hoja
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str)
            df = df.fillna('')
            
            self.data = df.values.tolist()
            return len(self.data), sheet_name
            
        except Exception as e:
            raise ValueError(f"Error al leer el archivo con pandas: {str(e)}")
    
    def parse_data(self, progress_callback=None, only_first_family=False):
        """Parsea los datos cargados y genera la estructura JSON.
        
        Args:
            progress_callback: Función de callback para mostrar progreso.
            only_first_family: Si es True, solo procesa la primera familia y sus subfamilias.
        """
        if self.data is None or len(self.data) == 0:
            raise ValueError("No hay datos cargados. Primero carga un archivo Excel.")
        
        # Reiniciar resultado
        self.result['familias'] = []
        self.result['articulos'] = []
        self.result['opcionales'] = []
        self.result['relaciones_articulo_opcional'] = []
        
        familia_actual = None
        subfamilia_actual = None
        en_seccion_opcional = False
        articulos_para_opcionales = []
        
        familias_dict = {}  # Para evitar duplicados (clave: capítulo)
        opcionales_dict = {}  # Para evitar duplicados de opcionales
        primera_familia_encontrada = False  # Para modo prueba
        segunda_familia_encontrada = False  # Para modo prueba
        
        # Reiniciar mapeo de capítulo a código hash
        self.capitulo_to_codigo = {}
        
        total_rows = len(self.data)
        
        for index, row in enumerate(self.data):
            if index == 0:  # Saltar encabezado
                continue
            
            if progress_callback:
                progress_callback(index, total_rows)
            
            # Extraer columnas según estructura del Excel:
            # A(0): ID | B(1): SKU | C(2): Código | D(3): Descripción | E(4): Description | 
            # F(5): Biblioteca Online | G(6): vacía | H(7): vacía | I(8): PVP 2025 | J(9): NUEVO
            id_col = str(row[0]).strip() if len(row) > 0 else ''
            sku = str(row[1]).strip() if len(row) > 1 else ''
            codigo = str(row[2]).strip() if len(row) > 2 else ''
            descripcion_es = str(row[3]).strip() if len(row) > 3 else ''
            descripcion_en = str(row[4]).strip() if len(row) > 4 else ''
            biblioteca_online = str(row[5]).strip() if len(row) > 5 else ''
            pvp = str(row[8]).strip() if len(row) > 8 else ''  # Columna I (índice 8)
            
            # Combinar todos los textos de la fila para buscar patrones
            all_texts = [id_col, sku, codigo, descripcion_es, descripcion_en]
            row_text = ' '.join([t for t in all_texts if t])
            
            # Ignorar filas completamente vacías
            if not row_text.strip() and not pvp:
                continue
            
            # ============================================
            # 1. DETECTAR SECCIÓN OPCIONAL (en cualquier columna)
            # ============================================
            es_fila_opcional = False
            for texto in all_texts:
                if texto.upper().strip() == 'OPCIONAL':
                    es_fila_opcional = True
                    break
            
            if es_fila_opcional:
                en_seccion_opcional = True
                continue
            
            # ============================================
            # 2. DETECTAR CATEGORÍA (formato: "N. NOMBRE")
            # ============================================
            categoria_encontrada = False
            for texto in all_texts:
                if not texto:
                    continue
                # Patrón: "1. ENFRIADORES" o "1 ENFRIADORES"
                match_cat = re.match(r'^(\d+)\.\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&]+)$', texto.strip())
                if not match_cat:
                    match_cat = re.match(r'^(\d+)\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&]+)$', texto.strip())
                
                if match_cat:
                    num_cat = match_cat.group(1)
                    nombre_cat = match_cat.group(2).strip()
                    
                    capitulo = num_cat  # Ej: "1", "2", "3"
                    
                    if capitulo not in familias_dict:
                        # Generar código hash de 8 caracteres
                        familia_codigo = self._generar_codfamilia(capitulo, nombre_cat)
                        self.capitulo_to_codigo[capitulo] = familia_codigo
                        
                        familias_dict[capitulo] = {
                            'codfamilia': familia_codigo,
                            'capitulo': capitulo,
                            'descripcion': nombre_cat,
                            'madre': None,
                            'nivel': 1
                        }
                    else:
                        # Actualizar nombre si estaba como placeholder
                        if familias_dict[capitulo]['descripcion'].startswith('Categoría '):
                            familias_dict[capitulo]['descripcion'] = nombre_cat
                            # Regenerar el código con el nombre correcto
                            familia_codigo = self._generar_codfamilia(capitulo, nombre_cat)
                            self.capitulo_to_codigo[capitulo] = familia_codigo
                            familias_dict[capitulo]['codfamilia'] = familia_codigo
                    
                    familia_actual = capitulo  # Guardamos el capítulo para referencia interna
                    subfamilia_actual = None
                    en_seccion_opcional = False
                    articulos_para_opcionales = []
                    categoria_encontrada = True
                    
                    # Control para modo prueba (solo primera familia)
                    if only_first_family:
                        if not primera_familia_encontrada:
                            primera_familia_encontrada = True
                        else:
                            segunda_familia_encontrada = True
                    break
            
            if categoria_encontrada:
                # Si estamos en modo prueba y ya encontramos la segunda familia, terminar
                if only_first_family and segunda_familia_encontrada:
                    break
                continue
            
            # ============================================
            # 3. DETECTAR SUBCATEGORÍA (formato: "N.N Nombre")
            # ============================================
            subcategoria_encontrada = False
            for texto in all_texts:
                if not texto:
                    continue
                # Patrón: "1.1 Refrigerador de botellas"
                match_sub = re.match(r'^(\d+)\.(\d+)\s+(.+)$', texto.strip())
                if match_sub:
                    num_cat = match_sub.group(1)
                    num_sub = match_sub.group(2)
                    nombre_sub = match_sub.group(3).strip()
                    
                    capitulo_madre = num_cat  # Ej: "1"
                    capitulo = f'{num_cat}.{num_sub}'  # Ej: "1.1"
                    
                    # Asegurar que existe la categoría padre
                    if capitulo_madre not in familias_dict:
                        # Generar UUID para la familia padre
                        familia_uuid = str(uuid.uuid4())
                        self.capitulo_to_uuid[capitulo_madre] = familia_uuid
                        
                        familias_dict[capitulo_madre] = {
                            'codfamilia': familia_uuid,
                            'capitulo': capitulo_madre,
                            'descripcion': f'Categoría {num_cat}',
                            'madre': None,
                            'nivel': 1
                        }
                    
                    # Crear subcategoría
                    if capitulo not in familias_dict:
                        # Generar UUID para esta subfamilia
                        subfamilia_uuid = str(uuid.uuid4())
                        self.capitulo_to_uuid[capitulo] = subfamilia_uuid
                        
                        # Obtener el UUID de la madre
                        madre_uuid = self.capitulo_to_uuid.get(capitulo_madre)
                        
                        familias_dict[capitulo] = {
                            'codfamilia': subfamilia_uuid,
                            'capitulo': capitulo,
                            'descripcion': nombre_sub,
                            'madre': madre_uuid,
                            'nivel': 2
                        }
                    
                    subfamilia_actual = capitulo  # Guardamos el capítulo para referencia interna
                    en_seccion_opcional = False
                    articulos_para_opcionales = []
                    subcategoria_encontrada = True
                    break
            
            if subcategoria_encontrada:
                continue
            
            # ============================================
            # 4. DETECTAR OPCIONAL (sin SKU, después de fila OPCIONAL)
            # ============================================
            # Un opcional NO tiene SKU (columna A vacía)
            if not sku and en_seccion_opcional:
                # La descripción del opcional está en Código (columna B) o Descripción (columna C)
                desc_opcional = codigo if codigo else descripcion_es
                desc_opcional_en = descripcion_en if descripcion_en else ''
                
                if desc_opcional:
                    self._procesar_opcional(desc_opcional, desc_opcional_en, pvp, 
                                           articulos_para_opcionales, opcionales_dict)
                continue
            
            # ============================================
            # 5. ES UN ARTÍCULO (tiene ID válido en columna A)
            # ============================================
            # Si tiene ID numérico y no es patrón de categoría, es un artículo
            if id_col and id_col.isdigit():
                en_seccion_opcional = False
                
                # Obtener el capítulo actual (subfamilia o familia)
                capitulo_actual = subfamilia_actual if subfamilia_actual else familia_actual
                # Convertir capítulo a UUID
                codfamilia_articulo = self.capitulo_to_uuid.get(capitulo_actual) if capitulo_actual else None
                precio = self._parsear_precio(pvp)
                
                articulo = {
                    'referencia': sku,  # SKU como referencia (ej: EFP1000EG)
                    'id': id_col,  # ID numérico (ej: 7128)
                    'nombre': codigo,  # Código legible (ej: EFP 1000 EG)
                    'codfamilia': codfamilia_articulo,
                    'descripcion_es': descripcion_es,
                    'descripcion_en': descripcion_en,
                    'precio': precio,
                    'activo': True
                }
                
                self.result['articulos'].append(articulo)
                articulos_para_opcionales.append(sku)
        
        # Convertir diccionarios a listas
        self.result['familias'] = list(familias_dict.values())
        
        # Actualizar metadata
        self.result['metadata']['total_categories'] = len([f for f in self.result['familias'] if f['nivel'] == 1])
        self.result['metadata']['total_subcategories'] = len([f for f in self.result['familias'] if f['nivel'] == 2])
        self.result['metadata']['total_articles'] = len(self.result['articulos'])
        self.result['metadata']['total_optionals'] = len(self.result['opcionales'])
        
        return self.result
    
    def _procesar_opcional(self, descripcion_es, descripcion_en, pvp, articulos, opcionales_dict):
        """Procesa una fila de opcional.
        
        Usa la descripción como clave para detectar duplicados dentro del mismo contexto,
        pero genera un UUID único para cada opcional para evitar colisiones entre
        opcionales de diferentes familias con la misma descripción.
        """
        # Usar la descripción como clave para detectar duplicados en el mismo contexto
        clave_duplicado = descripcion_es.lower().strip()
        
        if clave_duplicado not in opcionales_dict:
            # Generar UUID único para este opcional
            opcional_id = self._generar_id_opcional(descripcion_es)
            
            precio, precio_texto = self._parsear_precio_opcional(pvp)
            
            desc_final_es = descripcion_es
            desc_final_en = descripcion_en
            
            # Si hay texto de precio, añadirlo a la descripción
            if precio_texto:
                desc_final_es = f"{descripcion_es} ({precio_texto})"
                if descripcion_en:
                    desc_final_en = f"{descripcion_en} ({precio_texto})"
            
            opcional = {
                'id': opcional_id,
                'descripcion_es': desc_final_es,
                'descripcion_en': desc_final_en,
                'precio': precio
            }
            
            opcionales_dict[clave_duplicado] = opcional
            self.result['opcionales'].append(opcional)
        else:
            # Recuperar el ID del opcional ya existente
            opcional_id = opcionales_dict[clave_duplicado]['id']
        
        # Crear relaciones con los artículos
        for referencia in articulos:
            relacion = {
                'referencia': referencia,
                'opcional_id': opcional_id
            }
            self.result['relaciones_articulo_opcional'].append(relacion)
    
    def _generar_id_opcional(self, descripcion):
        """Genera un ID único para un opcional usando UUID.
        
        Se usa UUID para garantizar unicidad incluso cuando opcionales de 
        diferentes familias tienen la misma descripción.
        """
        return str(uuid.uuid4())
    
    def _generar_codfamilia(self, capitulo, descripcion):
        """Genera un código único de 8 caracteres basado en capítulo y descripción.
        
        Usa MD5 para crear un hash y toma los primeros 8 caracteres en mayúsculas.
        Esto garantiza que el mismo capítulo+descripción siempre genera el mismo código.
        """
        texto = f"{capitulo}:{descripcion}".lower()
        hash_md5 = hashlib.md5(texto.encode('utf-8')).hexdigest()
        return hash_md5[:8].upper()
    
    def _parsear_precio(self, valor):
        """Convierte un valor de precio a float."""
        if not valor:
            return 0.0
        
        # Limpiar el valor
        valor = str(valor).strip()
        
        # Remover símbolos de moneda y espacios
        valor = re.sub(r'[€$\s]', '', valor)
        
        # Manejar formato europeo (1.234,56) vs americano (1,234.56)
        if ',' in valor and '.' in valor:
            if valor.rfind(',') > valor.rfind('.'):
                # Formato europeo
                valor = valor.replace('.', '').replace(',', '.')
            else:
                # Formato americano
                valor = valor.replace(',', '')
        elif ',' in valor:
            valor = valor.replace(',', '.')
        
        try:
            return float(valor)
        except ValueError:
            return 0.0
    
    def _parsear_precio_opcional(self, valor):
        """
        Parsea el precio de un opcional.
        Retorna (precio, texto_adicional)
        - Si es "Sin incremento" o vacío -> (0.0, None)
        - Si es numérico -> (valor, None)
        - Si es otro texto -> (None, texto)
        """
        if not valor:
            return 0.0, None
        
        valor = str(valor).strip()
        valor_upper = valor.upper()
        
        # Sin incremento
        if 'SIN INCREMENTO' in valor_upper or 'NO INCREMENT' in valor_upper:
            return 0.0, None
        
        # Intentar parsear como número
        precio = self._parsear_precio(valor)
        if precio > 0:
            return precio, None
        
        # Es texto que no se puede convertir
        return None, valor
    
    def save_json(self, file_path):
        """Guarda el resultado en un archivo JSON."""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(self.result, f, ensure_ascii=False, indent=2)
        return file_path


class TarifaConverterApp:
    """Aplicación GUI con Tkinter para convertir tarifas."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor de Tarifa Excel a JSON - FSFramework")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        self.converter = TarifaConverter()
        self.current_file = None
        
        self._setup_styles()
        self._create_widgets()
        
    def _setup_styles(self):
        """Configura los estilos de la aplicación."""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Colores personalizados
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground='#2c3e50')
        style.configure('Subtitle.TLabel', font=('Segoe UI', 10), foreground='#7f8c8d')
        style.configure('Success.TLabel', foreground='#27ae60')
        style.configure('Error.TLabel', foreground='#e74c3c')
        style.configure('Info.TLabel', foreground='#3498db')
        
        style.configure('Action.TButton', font=('Segoe UI', 10, 'bold'), padding=10)
        style.configure('Secondary.TButton', font=('Segoe UI', 9), padding=5)
        
    def _create_widgets(self):
        """Crea todos los widgets de la interfaz."""
        # Frame principal con padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(title_frame, text="🔄 Conversor de Tarifa Excel a JSON", 
                  style='Title.TLabel').pack(anchor=tk.W)
        ttk.Label(title_frame, text="Convierte archivos de tarifa Excel al formato JSON para importación", 
                  style='Subtitle.TLabel').pack(anchor=tk.W)
        
        # Separador
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        
        # Sección de archivo
        file_frame = ttk.LabelFrame(main_frame, text="📁 Archivo de entrada", padding=15)
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        file_row = ttk.Frame(file_frame)
        file_row.pack(fill=tk.X)
        
        self.file_entry = ttk.Entry(file_row, width=60, state='readonly')
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Button(file_row, text="📂 Seleccionar Excel", 
                   command=self.select_file, style='Action.TButton').pack(side=tk.LEFT)
        
        # Frame para información del archivo
        self.file_info_frame = ttk.Frame(file_frame)
        self.file_info_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.file_info_label = ttk.Label(self.file_info_frame, text="", style='Info.TLabel')
        self.file_info_label.pack(anchor=tk.W)
        
        # Barra de progreso
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                             maximum=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X)
        
        self.progress_label = ttk.Label(progress_frame, text="", style='Subtitle.TLabel')
        self.progress_label.pack(anchor=tk.W, pady=(5, 0))
        
        # Opciones
        options_frame = ttk.LabelFrame(main_frame, text="⚙️ Opciones", padding=10)
        options_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.test_mode_var = tk.BooleanVar(value=False)
        test_mode_check = ttk.Checkbutton(
            options_frame, 
            text="🧪 Modo prueba (solo primera familia y sus subfamilias)",
            variable=self.test_mode_var
        )
        test_mode_check.pack(anchor=tk.W)
        ttk.Label(options_frame, text="Útil para probar la importación sin procesar todos los productos",
                  style='Subtitle.TLabel').pack(anchor=tk.W, padx=(20, 0))
        
        # Botones de acción
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.convert_btn = ttk.Button(action_frame, text="⚡ Convertir a JSON", 
                                       command=self.convert_file, style='Action.TButton',
                                       state='disabled')
        self.convert_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.save_btn = ttk.Button(action_frame, text="💾 Guardar JSON", 
                                    command=self.save_json, style='Action.TButton',
                                    state='disabled')
        self.save_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(action_frame, text="🔄 Limpiar", 
                   command=self.clear_all, style='Secondary.TButton').pack(side=tk.RIGHT)
        
        # Resultados / Estadísticas
        stats_frame = ttk.LabelFrame(main_frame, text="📊 Estadísticas", padding=15)
        stats_frame.pack(fill=tk.X, pady=(0, 15))
        
        stats_grid = ttk.Frame(stats_frame)
        stats_grid.pack(fill=tk.X)
        
        # Grid de estadísticas
        self.stats_labels = {}
        stats_items = [
            ('categories', '📁 Categorías:', 0, 0),
            ('subcategories', '📂 Subcategorías:', 0, 1),
            ('articles', '📦 Artículos:', 1, 0),
            ('optionals', '⚙️ Opcionales:', 1, 1),
        ]
        
        for key, text, row, col in stats_items:
            frame = ttk.Frame(stats_grid)
            frame.grid(row=row, column=col, sticky=tk.W, padx=20, pady=5)
            ttk.Label(frame, text=text, font=('Segoe UI', 10)).pack(side=tk.LEFT)
            self.stats_labels[key] = ttk.Label(frame, text="0", font=('Segoe UI', 10, 'bold'))
            self.stats_labels[key].pack(side=tk.LEFT, padx=(5, 0))
        
        # Log / Vista previa
        log_frame = ttk.LabelFrame(main_frame, text="📋 Vista previa / Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, font=('Consolas', 9),
                                                   wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Mensaje inicial
        self.log("Bienvenido al Conversor de Tarifa Excel a JSON")
        self.log("=" * 50)
        self.log("1. Selecciona un archivo Excel con la tarifa")
        self.log("2. Haz clic en 'Convertir a JSON'")
        self.log("3. Revisa las estadísticas y guarda el archivo JSON")
        self.log("")
        self.log("Estructura esperada del Excel:")
        self.log("  - Hoja: 'TARIFA COMPLETA'")
        self.log("  - Columnas: SKU | Código | Descripción | Description | PVP")
        
    def log(self, message, level='info'):
        """Añade un mensaje al log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        prefix = {'info': 'ℹ️', 'success': '✅', 'error': '❌', 'warning': '⚠️'}.get(level, 'ℹ️')
        
        self.log_text.insert(tk.END, f"[{timestamp}] {prefix} {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def select_file(self):
        """Abre el diálogo para seleccionar un archivo Excel."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo de tarifa",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Excel 2007+", "*.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if file_path:
            self.current_file = file_path
            self.file_entry.config(state='normal')
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.file_entry.config(state='readonly')
            
            self.log(f"Archivo seleccionado: {os.path.basename(file_path)}")
            
            # Intentar cargar el archivo
            try:
                self.progress_label.config(text="Cargando archivo...")
                self.root.update_idletasks()
                
                rows, sheet_name = self.converter.load_excel(file_path)
                
                self.file_info_label.config(
                    text=f"✅ Hoja encontrada: '{sheet_name}' | {rows:,} filas"
                )
                self.log(f"Archivo cargado correctamente: {rows:,} filas en hoja '{sheet_name}'", 'success')
                self.convert_btn.config(state='normal')
                self.progress_label.config(text="Listo para convertir")
                
            except Exception as e:
                self.file_info_label.config(text=f"❌ Error: {str(e)}")
                self.log(f"Error al cargar archivo: {str(e)}", 'error')
                self.convert_btn.config(state='disabled')
                self.progress_label.config(text="")
                
    def convert_file(self):
        """Convierte el archivo cargado a JSON."""
        if not self.converter.data:
            messagebox.showerror("Error", "Primero debes cargar un archivo Excel")
            return
        
        self.log("Iniciando conversión...", 'info')
        self.progress_var.set(0)
        self.convert_btn.config(state='disabled')
        
        try:
            def progress_callback(current, total):
                progress = (current / total) * 100
                self.progress_var.set(progress)
                self.progress_label.config(text=f"Procesando fila {current:,} de {total:,}...")
                if current % 100 == 0:
                    self.root.update_idletasks()
            
            result = self.converter.parse_data(progress_callback, only_first_family=self.test_mode_var.get())
            
            self.progress_var.set(100)
            self.progress_label.config(text="¡Conversión completada!")
            
            # Actualizar estadísticas
            meta = result['metadata']
            self.stats_labels['categories'].config(text=str(meta['total_categories']))
            self.stats_labels['subcategories'].config(text=str(meta['total_subcategories']))
            self.stats_labels['articles'].config(text=str(meta['total_articles']))
            self.stats_labels['optionals'].config(text=str(meta['total_optionals']))
            
            self.log(f"Conversión completada:", 'success')
            self.log(f"  - {meta['total_categories']} categorías")
            self.log(f"  - {meta['total_subcategories']} subcategorías")
            self.log(f"  - {meta['total_articles']} artículos")
            self.log(f"  - {meta['total_optionals']} opcionales")
            self.log(f"  - {len(result['relaciones_articulo_opcional'])} relaciones artículo-opcional")
            
            # Mostrar vista previa
            self.log("")
            self.log("Vista previa de familias (primeras 5):")
            for fam in result['familias'][:5]:
                self.log(f"  [Cap. {fam['capitulo']}] {fam['descripcion']} (nivel: {fam['nivel']})")
            
            self.log("")
            self.log("Vista previa de artículos (primeros 5):")
            for art in result['articulos'][:5]:
                self.log(f"  [{art['referencia']}] {art['nombre']} - €{art['precio']}")
            
            self.save_btn.config(state='normal')
            
        except Exception as e:
            self.log(f"Error durante la conversión: {str(e)}", 'error')
            messagebox.showerror("Error de conversión", str(e))
            
        finally:
            self.convert_btn.config(state='normal')
            
    def save_json(self):
        """Guarda el resultado en un archivo JSON."""
        if not self.converter.result['articulos']:
            messagebox.showerror("Error", "Primero debes convertir un archivo")
            return
        
        # Sugerir nombre basado en el archivo original
        suggested_name = "tarifa_completa.json"
        if self.current_file:
            base_name = os.path.splitext(os.path.basename(self.current_file))[0]
            suggested_name = f"{base_name}.json"
        
        file_path = filedialog.asksaveasfilename(
            title="Guardar archivo JSON",
            defaultextension=".json",
            initialfile=suggested_name,
            filetypes=[
                ("Archivo JSON", "*.json"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if file_path:
            try:
                self.converter.save_json(file_path)
                self.log(f"Archivo guardado: {file_path}", 'success')
                messagebox.showinfo("Éxito", f"Archivo JSON guardado correctamente:\n{file_path}")
            except Exception as e:
                self.log(f"Error al guardar: {str(e)}", 'error')
                messagebox.showerror("Error", f"Error al guardar el archivo:\n{str(e)}")
                
    def clear_all(self):
        """Limpia todo y reinicia la aplicación."""
        self.converter = TarifaConverter()
        self.current_file = None
        
        self.file_entry.config(state='normal')
        self.file_entry.delete(0, tk.END)
        self.file_entry.config(state='readonly')
        
        self.file_info_label.config(text="")
        self.progress_var.set(0)
        self.progress_label.config(text="")
        
        for label in self.stats_labels.values():
            label.config(text="0")
        
        self.convert_btn.config(state='disabled')
        self.save_btn.config(state='disabled')
        
        self.log_text.delete(1.0, tk.END)
        self.log("Aplicación reiniciada", 'info')


def main():
    """Punto de entrada principal."""
    # Verificar dependencias
    if openpyxl is None and pd is None:
        import sys
        print("Error: Se requiere openpyxl o pandas")
        print("Instala con: pip install openpyxl pandas")
        sys.exit(1)
    
    root = tk.Tk()
    app = TarifaConverterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
